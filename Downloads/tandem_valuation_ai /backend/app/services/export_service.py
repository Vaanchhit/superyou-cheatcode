import io
import os
from pathlib import Path
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXPORT_DIR = './backend/exports'
Path(EXPORT_DIR).mkdir(parents=True, exist_ok=True)


def build_workbook(payload: Dict[str, Any]) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    thin = Side(border_style='thin', color='DDDDDD')

    for sheet_data in payload.get('workbook', []):
        sheet = workbook.create_sheet(sheet_data['name'])
        sheet.sheet_view.zoomScale = 100
        for cell_data in sheet_data.get('cells', []):
            cell = sheet.cell(row=cell_data['row'], column=cell_data['col'])
            label = cell_data.get('label')
            if label and cell_data['col'] == 1:
                cell.value = label
            if cell_data.get('formula'):
                cell.value = cell_data['formula']
                cell.font = Font(color='0000AA')
            elif cell_data.get('value') is not None:
                cell.value = cell_data['value']
            if cell_data.get('editable'):
                cell.fill = PatternFill('solid', fgColor='FFF8DC')
            else:
                cell.fill = PatternFill('solid', fgColor='E8F1FF')
            cell.alignment = Alignment(vertical='center', horizontal='left')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 18

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_excel(payload: Dict[str, Any], filename: str) -> str:
    data = build_workbook(payload)
    output_path = os.path.join(EXPORT_DIR, filename)
    with open(output_path, 'wb') as file:
        file.write(data)
    return output_path
